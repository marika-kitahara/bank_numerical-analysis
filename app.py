import json
import re
import time
from io import BytesIO

import numpy as np
import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    from streamlit_local_storage import LocalStorage
except ImportError:
    LocalStorage = None

st.set_page_config(page_title="後方数値データ分析", layout="wide")
st.title("📊 後方数値データ分析ダッシュボード")


st.markdown(
    """
    <style>
    section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] {
        gap: 0.55rem;
    }
    section[data-testid="stSidebar"] label p {
        font-size: 0.88rem;
        font-weight: 600;
    }
    div[data-baseweb="select"] > div {
        min-height: 42px;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.4rem;
        flex-wrap: wrap;
    }
    .stTabs [data-baseweb="tab"] {
        height: 42px;
        padding-left: 0.8rem;
        padding-right: 0.8rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)



# ======================
# ブラウザ別設定保存
# ======================
PREFERENCE_STORAGE_KEY = "rear_numeric_analysis_preferences_v4"
PREFERENCE_SCHEMA_VERSION = 4
PREFERENCE_KEYS = [
    "data_sheet", "master_sheet", "use_cost_sheet", "cost_sheet",
    "filter_date_range", "filter_big", "filter_middle", "filter_small",
    "media_groups", "media_metrics", "media_trend_group", "media_trend_metric",
    "win_selected_middle", "win_pattern_cols", "win_target_metric", "win_min_count",
    "alloc_from_middle", "alloc_from_small", "alloc_to_middle", "alloc_to_small", "alloc_amount",
    "cross_x", "cross_y", "cross_metric", "cross_top_n",
    "seg_col", "seg_metric",
    "mail_smalls", "mail_campaigns", "mail_metrics", "mail_metric1", "mail_metric2",
    "mail_segment_col", "mail_segment_metric",
]


# 管理者デフォルト。ブラウザ保存値がない初回表示・設定初期化後はこの値を使う。
ADMIN_DEFAULTS = {
    "media_groups": ["大項目", "中項目", "小項目"],
    "media_trend_group": "中項目",
    "win_selected_middle": "すべて",
    "win_pattern_cols": ["小項目", "利用目的", "年齢グループ"],
    "win_target_metric": "承認率",
    "win_min_count": 30,
    "alloc_amount": 500000.0,
    "cross_x": "年齢グループ",
    "cross_y": "利用目的",
    "seg_col": "小項目",
    "mail_metrics": ["申込件数"],
    "mail_metric1": "申込件数",
    "mail_metric2": "承認率",
}


def admin_default(key, fallback=None):
    value = ADMIN_DEFAULTS.get(key, fallback)
    return list(value) if isinstance(value, list) else value




def ensure_widget_default(key, default, options=None, multiple=False):
    """ウィジェット生成前に管理者デフォルトを確実に反映する。

    ブラウザ保存値は consume_saved_preference() が後から上書きする。
    現在のセッションに古い値・候補外の値が残っている場合も安全に補正する。
    """
    valid = list(options) if options is not None else None

    if multiple:
        default_values = list(default or [])
        if valid is not None:
            default_values = [v for v in default_values if v in valid]

        current = st.session_state.get(key)
        if not isinstance(current, list):
            st.session_state[key] = default_values
        elif valid is not None:
            cleaned = [v for v in current if v in valid]
            st.session_state[key] = cleaned if cleaned else default_values
        return

    current = st.session_state.get(key)
    if current is None or (valid is not None and current not in valid):
        if valid is None:
            st.session_state[key] = default
        elif default in valid:
            st.session_state[key] = default
        elif valid:
            st.session_state[key] = valid[0]


def initialize_restored_defaults_once():
    """今回指定された管理者デフォルトを、古いセッション状態より優先して一度だけ復元する。"""
    marker = "_restored_admin_defaults_v4"
    if st.session_state.get(marker):
        return

    for key in [
        "media_groups", "media_trend_group", "win_pattern_cols",
        "alloc_amount", "cross_x", "cross_y", "seg_col",
        "mail_metrics", "mail_metric1", "mail_metric2",
    ]:
        st.session_state.pop(key, None)

    st.session_state[marker] = True


def serializable_value(value):
    """session_stateの値をlocalStorageへ保存できる形へ変換する。"""
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(value).isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, tuple):
        return [serializable_value(v) for v in value]
    if isinstance(value, list):
        return [serializable_value(v) for v in value]
    if isinstance(value, np.generic):
        return value.item()
    return value


def consume_saved_preference(
    key,
    options=None,
    default=None,
    multiple=False,
    converter=None,
    allow_empty_saved=False,
):
    """コード側デフォルトを土台にし、保存済みの有効な値だけを上書きする。

    - 保存値がない場合は、各ウィジェットに指定したデフォルトをそのまま使う。
    - 保存値が候補外・壊れている・空になってはいけない項目で空の場合は、デフォルトへ戻す。
    - フィルタのように空欄＝すべてを意味する項目は allow_empty_saved=True で空を許可できる。
    """
    pending = st.session_state.get("_pending_browser_preferences", {})
    if key not in pending:
        return

    value = pending.pop(key)
    st.session_state["_pending_browser_preferences"] = pending

    if converter is not None:
        try:
            value = converter(value)
        except Exception:
            value = default

    if options is not None:
        valid = list(options)
        if multiple:
            values = value if isinstance(value, list) else ([] if value is None else [value])
            value = [v for v in values if v in valid]
            if not value and not allow_empty_saved:
                value = list(default) if isinstance(default, (list, tuple)) else ([] if default is None else [default])
        else:
            if value not in valid:
                value = default if default in valid else (valid[0] if valid else None)
    elif multiple and value in (None, ""):
        value = [] if allow_empty_saved else (list(default) if isinstance(default, (list, tuple)) else default)

    if value is None and default is not None:
        value = default

    if value is not None:
        st.session_state[key] = value


def storage_get_item(storage, item_key, component_key=None):
    """streamlit-local-storage の版差を吸収して値を取得する。"""
    if storage is None:
        return None
    try:
        if component_key:
            return storage.getItem(item_key, key=component_key)
    except TypeError:
        # 古い版は Streamlit component 用の key 引数を受け付けない
        pass
    return storage.getItem(item_key)


def storage_set_item(storage, item_key, value, component_key=None):
    """streamlit-local-storage の版差を吸収して値を保存する。"""
    if storage is None:
        return None
    try:
        if component_key:
            return storage.setItem(item_key, value, key=component_key)
    except TypeError:
        # 古い版は Streamlit component 用の key 引数を受け付けない
        pass
    return storage.setItem(item_key, value)


def load_browser_preferences(storage):
    """localStorageから保存設定を読み込み、段階的適用用の待機領域へ入れる。"""
    if storage is None or st.session_state.get("_browser_preferences_loaded"):
        return

    try:
        raw = storage_get_item(
            storage,
            PREFERENCE_STORAGE_KEY,
            component_key="browser_preferences_raw",
        )
    except Exception:
        # localStorageコンポーネントの一時的な初期化失敗でアプリ全体を落とさない
        st.session_state["_browser_preferences_loaded"] = True
        return

    if raw in (None, "", {}):
        st.session_state["_browser_preferences_loaded"] = True
        return

    try:
        data = json.loads(raw) if isinstance(raw, str) else raw
        if isinstance(data, dict):
            # v2以降は保存値をenvelope化。旧形式も念のため読み取れる。
            if "values" in data:
                version = data.get("schema_version")
                values = data.get("values", {})
                if version != PREFERENCE_SCHEMA_VERSION or not isinstance(values, dict):
                    values = {}
            else:
                values = data

            st.session_state["_pending_browser_preferences"] = values
            st.session_state["_browser_preferences_loaded"] = True
            st.rerun()
        else:
            st.session_state["_browser_preferences_loaded"] = True
    except (TypeError, ValueError, json.JSONDecodeError):
        st.session_state["_browser_preferences_loaded"] = True


def collect_preferences():
    result = {}
    for key in PREFERENCE_KEYS:
        if key in st.session_state:
            result[key] = serializable_value(st.session_state[key])
    return result


local_storage = LocalStorage() if LocalStorage is not None else None
load_browser_preferences(local_storage)
initialize_restored_defaults_once()


CORE_METRICS = [
    "申込件数", "承認件数", "成約件数", "承認率", "成約率",
    "コスト", "申込CPA", "承認CPA",
    "取扱金額_当月", "取扱金額_翌月", "取扱金額_翌々月", "投下倍率",
]
RATE_METRICS = {"承認率", "成約率", "申込CPA", "承認CPA", "投下倍率"}

# Excelのふりがな情報などが見出しに連結された場合も吸収する
COLUMN_ALIASES = {
    "媒体名": ["媒体名", "媒体名バイタイメイ"],
    "媒体": ["媒体", "媒体バイタイ"],
    "成約フラグ": ["成約フラグ", "成約フラグセイヤク"],
    "単価": ["単価", "単価タンカ"],
    "借入フラグ": ["借入フラグ", "借入フラグカリイレ"],
    "大項目": ["大項目", "大項目ダイコウモク"],
    "中項目": ["中項目", "中項目チュウコウモク"],
    "小項目": ["小項目", "小項目ショウコウモク"],
    "小項目2": ["小項目2", "小項目2ショウコウモク"],
    "代理店コード": ["代理店コード", "代理店コードダイリテン"],
    "配信日": ["配信日", "配信日ハイシンビ"],
}


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]
    rename = {}
    for canonical, candidates in COLUMN_ALIASES.items():
        if canonical in df.columns:
            continue
        for col in df.columns:
            normalized = re.sub(r"\s+", "", str(col))
            if normalized in candidates:
                rename[col] = canonical
                break
    return df.rename(columns=rename)


def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    seen = {}
    cols = []
    for col in df.columns:
        col = str(col)
        seen[col] = seen.get(col, 0) + 1
        cols.append(col if seen[col] == 1 else f"{col}_{seen[col]}")
    df = df.copy()
    df.columns = cols
    return df


REQUIRED_SHEETS = [
    "後方数値データ(加工版)",
    "媒体コードマスタver3",
    "コストデータ",
]


@st.cache_data(show_spinner="必要な3シートを読み込んでいます…")
def read_required_sheets(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    """規定の3シートだけを1回のExcel読み込みで取得する。

    追加シートが存在しても分析対象にはせず、読み込み対象を固定する。
    """
    book = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=REQUIRED_SHEETS,
        engine="openpyxl",
    )
    return {
        sheet_name: make_unique_columns(clean_columns(df))
        for sheet_name, df in book.items()
    }


def safe_div(numerator, denominator):
    if isinstance(denominator, pd.Series):
        return numerator.div(denominator.replace(0, np.nan))
    return np.nan if denominator in (0, None) or pd.isna(denominator) else numerator / denominator


def normalize_month(value) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    dt = pd.to_datetime(text, errors="coerce")
    if pd.isna(dt):
        m = re.search(r"(20\d{2})\D+(\d{1,2})", text)
        return f"{m.group(1)}-{int(m.group(2)):02d}" if m else None
    return dt.strftime("%Y-%m")


def prepare_cost_lookup(cost_df: pd.DataFrame) -> pd.DataFrame:
    cost_df = clean_columns(cost_df)
    if cost_df.empty or len(cost_df.columns) < 5:
        return pd.DataFrame(columns=["小項目2", "申込月", "補完元コスト"])

    key_col = "項目4" if "項目4" in cost_df.columns else cost_df.columns[3]
    month_cols = [c for c in cost_df.columns if normalize_month(c)]
    if not month_cols:
        return pd.DataFrame(columns=["小項目2", "申込月", "補完元コスト"])

    long_df = cost_df[[key_col] + month_cols].melt(
        id_vars=[key_col], var_name="月", value_name="補完元コスト"
    )
    long_df = long_df.rename(columns={key_col: "小項目2"})
    long_df["小項目2"] = long_df["小項目2"].astype(str).str.strip()
    long_df["申込月"] = long_df["月"].map(normalize_month)
    long_df["補完元コスト"] = pd.to_numeric(long_df["補完元コスト"], errors="coerce")
    return (
        long_df.dropna(subset=["申込月", "補完元コスト"])
        .groupby(["小項目2", "申込月"], as_index=False)["補完元コスト"].sum()
    )


def preprocess(raw: pd.DataFrame, master: pd.DataFrame, cost_df: pd.DataFrame | None) -> pd.DataFrame:
    df = clean_columns(raw)
    master = clean_columns(master)

    required = ["申込日", "媒体コード"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"後方数値データに必須列がありません: {', '.join(missing)}")
    if "媒体コード" not in master.columns:
        raise ValueError("媒体コードマスタに『媒体コード』列がありません。")

    for col in ["申込日", "承認日", "成約日"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    df["申込月"] = df["申込日"].dt.strftime("%Y-%m")

    # 後方データに既に付与済みの列があっても、マスタの最新分類を優先
    master_cols = [
        "媒体コード", "テンプレコード", "メニュー名", "大項目", "中項目",
        "小項目", "小項目2", "代理店コード", "配信日", "ポイント",
        "キャンペーン識別", "メニューコード", "コストフラグ",
    ]
    master_cols = [c for c in master_cols if c in master.columns]
    df["媒体コード"] = df["媒体コード"].astype(str).str.strip()
    master["媒体コード"] = master["媒体コード"].astype(str).str.strip()
    master_join = master[master_cols].drop_duplicates(subset=["媒体コード"], keep="last")
    overlap = [c for c in master_cols if c != "媒体コード" and c in df.columns]
    df = df.drop(columns=overlap, errors="ignore").merge(master_join, on="媒体コード", how="left")

    df["申込フラグ"] = 1
    df["承認フラグ"] = df.get("承認日", pd.Series(index=df.index, dtype="datetime64[ns]")).notna().astype(int)
    if "成約フラグ" in df.columns:
        existing = pd.to_numeric(df["成約フラグ"], errors="coerce")
        date_flag = df.get("成約日", pd.Series(index=df.index, dtype="datetime64[ns]")).notna().astype(int)
        df["成約フラグ"] = existing.fillna(date_flag).astype(int)
    else:
        df["成約フラグ"] = df.get("成約日", pd.Series(index=df.index, dtype="datetime64[ns]")).notna().astype(int)

    numeric_candidates = [
        "取扱金額_申込当月", "取扱金額_申込翌月末", "取扱金額_申込翌々月末",
        "コスト", "単価", "年齢", "年収", "同借希望額", "子供数",
        "住宅ローン返済月額", "他社借入件数", "他社借入残高", "承認枠", "成約枠",
    ]
    for col in numeric_candidates:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # AK列コストがエラー/空欄の行だけ、コストデータの媒体×月総額を申込行数で按分
    df["コスト_元データ"] = pd.to_numeric(df.get("コスト", np.nan), errors="coerce")
    df["コスト補完フラグ"] = 0
    if cost_df is not None and not cost_df.empty and "小項目2" in df.columns:
        lookup = prepare_cost_lookup(cost_df)
        if not lookup.empty:
            df["小項目2"] = df["小項目2"].astype(str).str.strip()
            counts = df.groupby(["小項目2", "申込月"], dropna=False).size().rename("媒体月申込件数").reset_index()
            df = df.merge(counts, on=["小項目2", "申込月"], how="left")
            df = df.merge(lookup, on=["小項目2", "申込月"], how="left")
            fallback = safe_div(df["補完元コスト"], df["媒体月申込件数"])
            invalid = df["コスト_元データ"].isna()
            df["コスト"] = df["コスト_元データ"].where(~invalid, fallback)
            df.loc[invalid & df["コスト"].notna(), "コスト補完フラグ"] = 1
        else:
            df["コスト"] = df["コスト_元データ"]
    else:
        df["コスト"] = df["コスト_元データ"]

    if "年齢" in df.columns:
        df["年齢グループ"] = pd.cut(
            df["年齢"], [-np.inf, 29, 39, 49, 59, np.inf],
            labels=["20代以下", "30代", "40代", "50代", "60代以上"]
        )
    if "年収" in df.columns:
        df["年収グループ"] = pd.cut(
            df["年収"], [-np.inf, 299, 499, 699, 999, np.inf],
            labels=["300万未満", "300-500万", "500-700万", "700-1000万", "1000万以上"]
        )
    return df


def aggregate(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    group_cols = list(dict.fromkeys([c for c in group_cols if c in df.columns]))
    work = df.copy()
    for col in ["取扱金額_申込当月", "取扱金額_申込翌月末", "取扱金額_申込翌々月末", "コスト"]:
        if col not in work.columns:
            work[col] = 0.0

    agg_map = {
        "申込件数": ("申込フラグ", "sum"),
        "承認件数": ("承認フラグ", "sum"),
        "成約件数": ("成約フラグ", "sum"),
        "コスト": ("コスト", "sum"),
        "取扱金額_当月": ("取扱金額_申込当月", "sum"),
        "取扱金額_翌月": ("取扱金額_申込翌月末", "sum"),
        "取扱金額_翌々月": ("取扱金額_申込翌々月末", "sum"),
    }
    result = work.groupby(group_cols, dropna=False, observed=True).agg(**agg_map).reset_index() if group_cols else pd.DataFrame({
        name: [work[src].agg(func)] for name, (src, func) in agg_map.items()
    })
    result["承認率"] = safe_div(result["承認件数"], result["申込件数"])
    result["成約率"] = safe_div(result["成約件数"], result["申込件数"])
    result["申込CPA"] = safe_div(result["コスト"], result["申込件数"])
    result["承認CPA"] = safe_div(result["コスト"], result["承認件数"])
    result["投下倍率"] = safe_div(result["取扱金額_翌月"], result["コスト"])
    return result


def add_extra_numeric_sums(df: pd.DataFrame, result: pd.DataFrame, group_cols: list[str], selected: list[str]) -> pd.DataFrame:
    # 集計軸が数値型として読み込まれていても、表示項目側へ重複追加しない。
    # 例：空欄が多い「大項目」がfloat型になった場合、reset_index時に列名衝突していた。
    group_cols = list(dict.fromkeys(c for c in group_cols if c in df.columns))
    selected = list(dict.fromkeys(selected))
    extra = [
        c for c in selected
        if c not in CORE_METRICS
        and c not in group_cols
        and c not in result.columns
        and c in df.columns
        and pd.api.types.is_numeric_dtype(df[c])
    ]
    if not extra:
        return result
    if group_cols:
        sums = (
            df.groupby(group_cols, dropna=False, observed=True)[extra]
            .sum(min_count=1)
            .reset_index()
        )
        return result.merge(sums, on=group_cols, how="left", validate="one_to_one")
    for col in extra:
        result[col] = df[col].sum(min_count=1)
    return result


def format_table(df: pd.DataFrame):
    formats = {c: "{:.1%}" for c in ["承認率", "成約率"] if c in df.columns}
    formats.update({c: "{:,.0f}" for c in ["コスト", "申込CPA", "承認CPA", "取扱金額_当月", "取扱金額_翌月", "取扱金額_翌々月"] if c in df.columns})
    formats.update({c: "{:.2f}" for c in ["投下倍率"] if c in df.columns})
    return df.style.format(formats, na_rep="-")


def to_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="分析結果")
    return output.getvalue()


def _excel_month_key(value) -> str | None:
    """Excelセル値を YYYY-MM に正規化する。"""
    if value in (None, ""):
        return None
    dt = pd.to_datetime(value, errors="coerce")
    if pd.notna(dt):
        return pd.Timestamp(dt).strftime("%Y-%m")
    text = str(value).strip()
    m = re.search(r"(20\d{2})\D*(\d{1,2})", text)
    return f"{m.group(1)}-{int(m.group(2)):02d}" if m else None


def _excel_key(value) -> str:
    """MATCH/COUNTIFS用に文字列キーを揃える。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@st.cache_data(show_spinner="AK列を計算したExcelを作成しています…")
def build_ak_export(file_bytes: bytes) -> bytes:
    """元ブックの『後方数値データ(加工版)』を保ったままAK列だけ計算して返す。

    計算仕様:
      numerator = INDEX(コストデータ!E:右端, MATCH(AF行, コストデータ!D:D), MATCH(B行の年月, 月見出し))
      denominator = 各行ごとに、後方数値データで B列の年月が同じ かつ C列=L行 かつ D列=M行 の件数
      AK = numerator / denominator

    コストデータの月列はE列以降を右端まで動的に探索する。
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=False)
    target_name = "後方数値データ(加工版)"
    cost_name = "コストデータ"
    if target_name not in wb.sheetnames or cost_name not in wb.sheetnames:
        raise ValueError("『後方数値データ(加工版)』または『コストデータ』シートがありません。")

    ws = wb[target_name]
    cost_ws = wb[cost_name]

    # コストデータ D列のキー → 行番号
    cost_row_by_key = {}
    for r in range(2, cost_ws.max_row + 1):
        key = _excel_key(cost_ws.cell(r, 4).value)
        if key:
            cost_row_by_key[key] = r

    # コストデータ E列以降の年月 → 列番号（V列より右に増えても対応）
    cost_col_by_month = {}
    for c in range(5, cost_ws.max_column + 1):
        month = _excel_month_key(cost_ws.cell(1, c).value)
        if month:
            cost_col_by_month[month] = c

    # 分母用に「年月 × C列 × D列」の件数を一度だけ集計する。
    # 各出力行では、その行自身の L列・M列を条件として参照する。
    denominator_counts = {}
    for r in range(2, ws.max_row + 1):
        month = _excel_month_key(ws.cell(r, 2).value)  # B列
        c_key = _excel_key(ws.cell(r, 3).value)        # C列
        d_key = _excel_key(ws.cell(r, 4).value)        # D列
        if not month:
            continue
        group_key = (month, c_key, d_key)
        denominator_counts[group_key] = denominator_counts.get(group_key, 0) + 1

    # AK列 = 37列目。元シートの他セルには触れない。
    for r in range(2, ws.max_row + 1):
        month = _excel_month_key(ws.cell(r, 2).value)  # B列
        af_key = _excel_key(ws.cell(r, 32).value)      # AF列
        criterion_c = _excel_key(ws.cell(r, 12).value)  # L列：各行
        criterion_d = _excel_key(ws.cell(r, 13).value)  # M列：各行
        denominator = denominator_counts.get((month, criterion_c, criterion_d), 0)
        cost_row = cost_row_by_key.get(af_key)
        cost_col = cost_col_by_month.get(month)

        result = None
        if denominator and cost_row and cost_col:
            numerator = cost_ws.cell(cost_row, cost_col).value
            numerator = pd.to_numeric(numerator, errors="coerce")
            if pd.notna(numerator):
                result = float(numerator) / denominator

        ws.cell(r, 37).value = result

    # エクスポート対象は指定シートのみ。対象シート自体のセル・書式は元ブックを利用。
    for sheet_name in list(wb.sheetnames):
        if sheet_name != target_name:
            del wb[sheet_name]

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def metric_options(df: pd.DataFrame) -> list[str]:
    # 分類・識別系の列は、Excel上で全空欄だと数値型(float)判定されることがある。
    # それらを表示指標に混ぜず、実際の数値項目だけを候補にする。
    excluded = {
        "申込フラグ", "承認フラグ", "成約フラグ", "コスト_元データ",
        "コスト補完フラグ", "媒体月申込件数", "補完元コスト",
        "媒体コード", "テンプレコード", "メニュー名", "大項目", "中項目",
        "小項目", "小項目2", "代理店コード", "配信日", "ポイント",
        "キャンペーン識別", "メニューコード", "コストフラグ", "媒体名", "媒体",
        "申込月",
    }
    extras = [
        c for c in df.select_dtypes(include="number").columns
        if c not in excluded and c not in CORE_METRICS
    ]
    return list(dict.fromkeys(CORE_METRICS + extras))


def selection_filter(df: pd.DataFrame, column: str, label: str, key: str) -> pd.DataFrame:
    """候補が存在する分類列だけをサイドバーへ表示する。"""
    if column not in df.columns:
        return df

    values = df[column].dropna().astype(str).str.strip()
    choices = sorted(v for v in values.unique().tolist() if v and v.casefold() != "nan")
    if not choices:
        st.sidebar.caption(f"{label}: 現在選択可能な値なし")
        return df

    consume_saved_preference(key, choices, default=[], multiple=True, allow_empty_saved=True)
    selected = st.sidebar.multiselect(
        label,
        choices,
        key=key,
        placeholder="すべて",
    )
    return df[df[column].astype(str).isin(selected)] if selected else df


def render_kpis(df: pd.DataFrame):
    kpi = aggregate(df, [])
    row = kpi.iloc[0]

    st.markdown(
        """
        <style>
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid rgba(49, 51, 63, 0.18);
            border-radius: 12px;
            padding: 14px 16px;
            min-height: 112px;
            box-shadow: 0 2px 7px rgba(0, 0, 0, 0.05);
        }
        div[data-testid="stMetricLabel"] p {
            font-size: 0.95rem !important;
            font-weight: 700 !important;
        }
        div[data-testid="stMetricValue"] {
            font-size: clamp(1.45rem, 2.1vw, 2.15rem) !important;
            line-height: 1.15 !important;
            overflow: visible !important;
            white-space: nowrap !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # 7枚を一列に詰め込むと金額が省略されるため、4枚＋3枚の二段表示にする。
    first = st.columns(4)
    first[0].metric("申込件数", f"{row['申込件数']:,.0f}")
    first[1].metric("承認件数", f"{row['承認件数']:,.0f}")
    first[2].metric("承認率", f"{row['承認率']:.1%}" if pd.notna(row["承認率"]) else "-")
    first[3].metric("コスト", f"¥{row['コスト']:,.0f}" if pd.notna(row["コスト"]) else "-")

    second = st.columns(3)
    second[0].metric("申込CPA", f"¥{row['申込CPA']:,.0f}" if pd.notna(row["申込CPA"]) else "-")
    second[1].metric("承認CPA", f"¥{row['承認CPA']:,.0f}" if pd.notna(row["承認CPA"]) else "-")
    second[2].metric("投下倍率", f"{row['投下倍率']:.2f}" if pd.notna(row["投下倍率"]) else "-")


def allocation_simulator(df: pd.DataFrame):
    st.subheader("💸 媒体アロケーション影響試算")
    st.caption("過去実績の媒体別効率が移管後も続く前提で、元媒体から減る成果と先媒体で増える成果の差分を試算します。")
    if not {"中項目", "小項目", "コスト"}.issubset(df.columns):
        st.info("中項目・小項目・コストが揃うと利用できます。")
        return

    def valid_options(series: pd.Series) -> list[str]:
        values = series.dropna().astype(str).str.strip()
        return sorted(v for v in values.unique().tolist() if v and v.casefold() != "nan")

    middles = valid_options(df["中項目"])
    if not middles:
        st.info("現在のフィルタ条件では、選択可能な中項目がありません。")
        return

    consume_saved_preference("alloc_from_middle", middles, default=middles[0])
    consume_saved_preference("alloc_to_middle", middles, default=middles[1] if len(middles) > 1 else middles[0])
    consume_saved_preference("alloc_amount", default=admin_default("alloc_amount", 500000.0))

    # 依存する選択肢が変わった際、古いsession_state値が残って選択不能になるのを防ぐ。
    if st.session_state.get("alloc_from_middle") not in middles:
        st.session_state["alloc_from_middle"] = middles[0]
    if st.session_state.get("alloc_to_middle") not in middles:
        st.session_state["alloc_to_middle"] = middles[1] if len(middles) > 1 else middles[0]

    c1, c2, c3 = st.columns(3)
    from_middle = c1.selectbox("アロケーション元：中項目", middles, key="alloc_from_middle")
    from_smalls = valid_options(df.loc[df["中項目"].astype(str).str.strip() == from_middle, "小項目"])
    if not from_smalls:
        c2.warning("元媒体の小項目がありません。")
        return
    consume_saved_preference("alloc_from_small", from_smalls, default=from_smalls[0])
    if st.session_state.get("alloc_from_small") not in from_smalls:
        st.session_state["alloc_from_small"] = from_smalls[0]
    from_small = c2.selectbox("アロケーション元：小項目", from_smalls, key="alloc_from_small")
    amount = c3.number_input(
        "移管金額（円）", min_value=0.0, value=500000.0,
        step=10000.0, format="%.0f", key="alloc_amount"
    )

    d1, d2 = st.columns(2)
    to_middle = d1.selectbox("アロケーション先：中項目", middles, key="alloc_to_middle")
    to_smalls = valid_options(df.loc[df["中項目"].astype(str).str.strip() == to_middle, "小項目"])
    if not to_smalls:
        d2.warning("移管先の小項目がありません。")
        return
    preferred_to = next((v for v in to_smalls if not (to_middle == from_middle and v == from_small)), to_smalls[0])
    consume_saved_preference("alloc_to_small", to_smalls, default=preferred_to)
    if st.session_state.get("alloc_to_small") not in to_smalls:
        st.session_state["alloc_to_small"] = preferred_to
    to_small = d2.selectbox("アロケーション先：小項目", to_smalls, key="alloc_to_small")

    same_media = from_middle == to_middle and from_small == to_small
    if same_media:
        st.warning("アロケーション元と先が同じです。異なる媒体を選択してください。")

    if st.button("影響を試算", type="primary", disabled=amount <= 0 or same_media):
        source = df[
            (df["中項目"].astype(str).str.strip() == from_middle)
            & (df["小項目"].astype(str).str.strip() == from_small)
        ]
        dest = df[
            (df["中項目"].astype(str).str.strip() == to_middle)
            & (df["小項目"].astype(str).str.strip() == to_small)
        ]
        if source.empty or dest.empty:
            st.error("元または先媒体の対象データがありません。")
            return

        s = aggregate(source, []).iloc[0]
        d = aggregate(dest, []).iloc[0]
        if s["コスト"] <= 0 or d["コスト"] <= 0:
            st.error("元または先媒体の実績コストが0のため試算できません。")
            return

        def per_cost(row, metric):
            return row[metric] / row["コスト"] if row["コスト"] else np.nan

        before = aggregate(pd.concat([source, dest]), []).iloc[0]
        source_cost_after = max(s["コスト"] - amount, 0)
        effective_amount = min(amount, s["コスト"])
        dest_cost_after = d["コスト"] + effective_amount

        projected = {}
        for metric in ["申込件数", "承認件数", "成約件数", "取扱金額_翌月"]:
            projected[metric] = per_cost(s, metric) * source_cost_after + per_cost(d, metric) * dest_cost_after
        projected["コスト"] = source_cost_after + dest_cost_after
        projected["承認率"] = safe_div(projected["承認件数"], projected["申込件数"])
        projected["申込CPA"] = safe_div(projected["コスト"], projected["申込件数"])
        projected["承認CPA"] = safe_div(projected["コスト"], projected["承認件数"])
        projected["投下倍率"] = safe_div(projected["取扱金額_翌月"], projected["コスト"])

        rows = []
        for metric in ["投下倍率", "取扱金額_翌月", "申込CPA", "承認CPA", "申込件数", "承認率"]:
            b = before[metric]
            a = projected[metric]
            rows.append({"指標": metric, "移管前": b, "移管後試算": a, "差分": a - b, "変化率": safe_div(a - b, b)})
        impact = pd.DataFrame(rows)
        st.dataframe(format_table(impact), width="stretch", hide_index=True)
        if effective_amount < amount:
            st.warning(f"元媒体の実績コストを超えるため、試算上は {effective_amount:,.0f}円で計算しました。")


uploaded_file = st.file_uploader(
    "後方数値データ・媒体コードマスタ・コストデータを含むExcel",
    type=["xlsx"],
)

if uploaded_file is None:
    st.info(
        "Excelファイルをアップロードしてください。"
        "追加シートが含まれていても、規定の3シートだけを読み込みます。"
    )
    st.stop()

file_bytes = uploaded_file.getvalue()

try:
    loaded_sheets = read_required_sheets(file_bytes)
    raw_df = loaded_sheets["後方数値データ(加工版)"]
    master_df = loaded_sheets["媒体コードマスタver3"]
    cost_df = loaded_sheets["コストデータ"]
    base_df = preprocess(raw_df, master_df, cost_df)
except ValueError as exc:
    # pandas/openpyxlが規定シート不足を検知した場合に、利用者へ分かりやすく案内する。
    message = str(exc)
    if "Worksheet named" in message:
        st.error(
            "規定シートが見つかりません。"
            "Excel内に次の3シートが存在することを確認してください："
            "「後方数値データ(加工版)」「媒体コードマスタver3」「コストデータ」"
        )
    else:
        st.error(f"読み込み・前処理でエラーが発生しました: {exc}")
    st.stop()
except Exception as exc:
    st.error(f"読み込み・前処理でエラーが発生しました: {exc}")
    st.stop()

st.sidebar.header("🔍 全タブ共通フィルタ")
df = base_df.copy()
valid_dates = df["申込日"].dropna()
if not valid_dates.empty:
    default_date_range = [valid_dates.min().date(), valid_dates.max().date()]
    def _date_converter(value):
        values = value if isinstance(value, list) else [value]
        return [pd.to_datetime(v).date() for v in values]
    consume_saved_preference("filter_date_range", default=default_date_range, converter=_date_converter)
    date_range = st.sidebar.date_input(
        "申込日（期間）", default_date_range, key="filter_date_range"
    )
    if len(date_range) == 2:
        df = df[(df["申込日"] >= pd.Timestamp(date_range[0])) & (df["申込日"] < pd.Timestamp(date_range[1]) + pd.Timedelta(days=1))]

df = selection_filter(df, "大項目", "大項目", "filter_big")
df = selection_filter(df, "中項目", "中項目", "filter_middle")
df = selection_filter(df, "小項目", "小項目", "filter_small")
st.sidebar.caption(f"対象件数: {len(df):,} / {len(base_df):,}")
if "コスト補完フラグ" in df.columns:
    st.sidebar.caption(f"コスト補完行: {int(df['コスト補完フラグ'].sum()):,}")

if df.empty:
    st.warning("フィルタ条件に該当するデータがありません。")
    st.stop()

render_kpis(df)

st.divider()
st.subheader("📤 AK列計算済みデータのエクスポート")
st.caption("元の『後方数値データ(加工版)』をそのまま使い、AK列だけ指定計算で更新したExcelを出力します。")
try:
    ak_export_bytes = build_ak_export(file_bytes)
    st.download_button(
        "📥 AK列計算済みの後方数値データをダウンロード",
        ak_export_bytes,
        "後方数値データ_AK計算済み.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_ak_export",
    )
except Exception as exc:
    st.error(f"AK列計算済みExcelの作成に失敗しました: {exc}")

# 分析軸候補：日付・識別・属性列。内部計算列は除外
internal_cols = {"申込フラグ", "承認フラグ", "成約フラグ", "コスト_元データ", "コスト補完フラグ", "媒体月申込件数", "補完元コスト"}
dimension_cols = [
    c for c in df.columns
    if c not in internal_cols and not pd.api.types.is_numeric_dtype(df[c]) and c not in ["承認日", "成約日"]
]
for preferred in ["申込月", "大項目", "中項目", "小項目", "小項目2", "媒体名", "キャンペーン識別", "メニューコード", "年齢グループ", "年収グループ"]:
    if preferred in df.columns and preferred not in dimension_cols:
        dimension_cols.insert(0, preferred)
dimension_cols = list(dict.fromkeys(dimension_cols))
all_metrics = metric_options(df)

tab_media, tab_win, tab_cross, tab_seg, tab_mail = st.tabs([
    "📊 媒体分析", "🏆 勝ちパターン", "🔄 クロス分析", "👤 セグメント別", "✉️ メルマガ"
])

with tab_media:
    st.subheader("媒体分析")
    default_groups = [c for c in ["大項目", "中項目", "小項目"] if c in dimension_cols]
    media_group_defaults = [c for c in admin_default("media_groups", default_groups) if c in dimension_cols]
    ensure_widget_default("media_groups", media_group_defaults, dimension_cols, multiple=True)
    consume_saved_preference("media_groups", dimension_cols, default=media_group_defaults, multiple=True)
    groups = st.multiselect("集計軸", dimension_cols, default=media_group_defaults, key="media_groups")
    default_metrics = [m for m in CORE_METRICS if m in all_metrics]
    metric_actions = st.columns([1, 1, 6])
    if metric_actions[0].button("全選択", key="media_select_all"):
        st.session_state["media_metrics"] = all_metrics
        st.rerun()
    if metric_actions[1].button("基本項目", key="media_select_core"):
        st.session_state["media_metrics"] = default_metrics
        st.rerun()
    consume_saved_preference("media_metrics", all_metrics, default=default_metrics, multiple=True)
    selected_metrics = st.multiselect("表示項目", all_metrics, default=default_metrics, key="media_metrics")
    res = aggregate(df, groups)
    res = add_extra_numeric_sums(df, res, groups, selected_metrics)
    display_cols = list(dict.fromkeys(groups + [m for m in selected_metrics if m in res.columns]))
    if not display_cols:
        st.info("集計軸または表示項目を1つ以上選択してください。")
    else:
        media_output = res[display_cols]
        st.dataframe(format_table(media_output), width="stretch", hide_index=True)
        st.download_button("📥 媒体分析をダウンロード", to_excel(media_output), "媒体分析.xlsx")

    if "申込日" in df.columns and groups:
        default_trend_group = "中項目" if "中項目" in groups else groups[0]
        trend_metric_options = [m for m in selected_metrics if m in CORE_METRICS]
        if "申込件数" not in trend_metric_options:
            trend_metric_options = ["申込件数"] + trend_metric_options

        ensure_widget_default("media_trend_group", default_trend_group, groups)
        ensure_widget_default("media_trend_metric", "申込件数", trend_metric_options)
        consume_saved_preference("media_trend_group", groups, default=default_trend_group)
        consume_saved_preference("media_trend_metric", trend_metric_options, default="申込件数")

        trend_group = st.selectbox("日次推移の系列", groups, key="media_trend_group")
        trend_metric = st.selectbox("日次推移の指標", trend_metric_options, key="media_trend_metric")

        daily_df = df.loc[df["申込日"].notna()].copy()
        daily_df["日付"] = daily_df["申込日"].dt.normalize()
        daily = aggregate(daily_df, ["日付", trend_group])
        pivot = (
            daily.pivot_table(index="日付", columns=trend_group, values=trend_metric, aggfunc="sum")
            .sort_index()
            .dropna(axis=1, how="all")
        )
        if pivot.empty:
            st.info("日次推移を表示できるデータがありません。")
        else:
            st.line_chart(pivot, width="stretch")

with tab_win:
    st.subheader("勝ちパターン分析")
    middle_options = sorted(df["中項目"].dropna().astype(str).str.strip().unique()) if "中項目" in df.columns else []
    middle_options = [v for v in middle_options if v and v.casefold() != "nan"]
    win_middle_options = ["すべて"] + middle_options
    consume_saved_preference("win_selected_middle", win_middle_options, default=admin_default("win_selected_middle", "すべて"))
    selected_middle = st.selectbox("分析する中項目", win_middle_options, key="win_selected_middle")
    win_df = df if selected_middle == "すべて" else df[df["中項目"].astype(str).str.strip() == selected_middle]
    pattern_options = [c for c in ["小項目", "媒体名", "キャンペーン識別", "年齢グループ", "年収グループ", "性別", "利用目的", "都道府県"] if c in dimension_cols]
    pattern_defaults = [c for c in admin_default("win_pattern_cols", ["小項目", "利用目的", "年齢グループ"]) if c in pattern_options]
    consume_saved_preference("win_pattern_cols", pattern_options, default=pattern_defaults, multiple=True)
    pattern_cols = st.multiselect("勝ちパターンの組み合わせ", pattern_options, default=pattern_defaults, key="win_pattern_cols")
    win_metric_options = ["承認率", "申込件数", "投下倍率", "取扱金額_翌月", "申込CPA", "承認CPA", "成約件数", "成約率"]
    consume_saved_preference("win_target_metric", win_metric_options, default=admin_default("win_target_metric", "承認率"))
    consume_saved_preference("win_min_count", default=admin_default("win_min_count", 30))
    target_metric = st.selectbox("最大化する指標", win_metric_options, key="win_target_metric")
    min_count = st.number_input("最低申込件数", min_value=1, value=int(admin_default("win_min_count", 30)), step=10, key="win_min_count")
    if pattern_cols:
        win = aggregate(win_df, pattern_cols)
        win = win[win["申込件数"] >= min_count]
        ascending = target_metric in {"申込CPA", "承認CPA"}
        win = win.sort_values(target_metric, ascending=ascending, na_position="last")
        st.dataframe(format_table(win.head(30)), width="stretch", hide_index=True)
        if not win.empty:
            top = win.iloc[0]
            combo = " × ".join(str(top[c]) for c in pattern_cols)
            direction = "最小" if ascending else "最大"
            st.success(f"{target_metric}が{direction}の勝ちパターン：{combo}")
        st.download_button("📥 勝ちパターンをダウンロード", to_excel(win), "勝ちパターン.xlsx")
    allocation_simulator(win_df)

with tab_cross:
    st.subheader("クロス分析")
    default_cross_x = admin_default("cross_x", "年齢グループ") if admin_default("cross_x", "年齢グループ") in dimension_cols else dimension_cols[0]
    default_cross_y = admin_default("cross_y", "利用目的") if admin_default("cross_y", "利用目的") in dimension_cols else dimension_cols[min(1, len(dimension_cols)-1)]
    cross_metric_options = [m for m in CORE_METRICS if m in all_metrics]
    ensure_widget_default("cross_x", default_cross_x, dimension_cols)
    ensure_widget_default("cross_y", default_cross_y, dimension_cols)
    ensure_widget_default("cross_metric", cross_metric_options[0], cross_metric_options)
    consume_saved_preference("cross_x", dimension_cols, default=default_cross_x)
    consume_saved_preference("cross_y", dimension_cols, default=default_cross_y)
    consume_saved_preference("cross_metric", cross_metric_options, default=cross_metric_options[0])
    consume_saved_preference("cross_top_n", default=15)
    x = st.selectbox("X軸", dimension_cols, key="cross_x")
    y = st.selectbox("Y軸", dimension_cols, key="cross_y")
    metric = st.selectbox("指標", cross_metric_options, key="cross_metric")
    if x != y:
        cross = aggregate(df, [x, y])
        st.dataframe(format_table(cross), width="stretch", hide_index=True)
        pivot = cross.pivot_table(index=x, columns=y, values=metric, aggfunc="sum")

        st.subheader("クロス集計表")
        st.caption("matplotlibに依存しない表示です。行・列をクリックして並べ替えできます。")
        if metric in {"承認率", "成約率"}:
            pivot_display = pivot.map(lambda v: f"{v:.1%}" if pd.notna(v) else "-")
        elif metric == "投下倍率":
            pivot_display = pivot.map(lambda v: f"{v:.2f}" if pd.notna(v) else "-")
        else:
            pivot_display = pivot.map(lambda v: f"{v:,.0f}" if pd.notna(v) else "-")
        st.dataframe(pivot_display, width="stretch")

        st.subheader("上位組み合わせ")
        sort_ascending = metric in {"申込CPA", "承認CPA"}
        top_n = st.slider("表示件数", min_value=5, max_value=50, value=15, step=5, key="cross_top_n")
        ranking_cols = [x, y, metric]
        ranking = cross[ranking_cols].sort_values(metric, ascending=sort_ascending, na_position="last").head(top_n)
        st.dataframe(format_table(ranking), width="stretch", hide_index=True)

        st.download_button("📥 クロス分析をダウンロード", to_excel(cross), "クロス分析.xlsx")

with tab_seg:
    st.subheader("セグメント別分析")
    default_seg = admin_default("seg_col", "小項目") if admin_default("seg_col", "小項目") in dimension_cols else dimension_cols[0]
    seg_metric_options = [m for m in CORE_METRICS if m in all_metrics]
    ensure_widget_default("seg_col", default_seg, dimension_cols)
    ensure_widget_default("seg_metric", seg_metric_options[0], seg_metric_options)
    consume_saved_preference("seg_col", dimension_cols, default=default_seg)
    consume_saved_preference("seg_metric", seg_metric_options, default=seg_metric_options[0])
    seg = st.selectbox("分析項目", dimension_cols, key="seg_col")
    chart_metric = st.selectbox("縦棒グラフの指標", seg_metric_options, key="seg_metric")
    seg_res = aggregate(df, [seg])
    toggle = seg_res[[seg]].copy()
    toggle.insert(0, "表示", True)
    edited = st.data_editor(
        toggle,
        hide_index=True,
        width="stretch",
        disabled=[seg],
        column_config={"表示": st.column_config.CheckboxColumn("グラフ表示")},
        key="seg_toggle",
    )
    selected_values = edited.loc[edited["表示"], seg].tolist()
    chart_df = seg_res[seg_res[seg].isin(selected_values)].set_index(seg)[[chart_metric]]
    st.bar_chart(chart_df)
    st.dataframe(format_table(seg_res), width="stretch", hide_index=True)
    st.download_button("📥 セグメント分析をダウンロード", to_excel(seg_res), "セグメント分析.xlsx")

with tab_mail:
    st.subheader("メルマガ分析（中項目 = Mail）")

    if "中項目" not in df.columns:
        st.info("中項目列がありません。")
    else:
        mail_df = df[df["中項目"].astype(str).str.strip().str.casefold() == "mail"].copy()

        if mail_df.empty:
            st.info("現在の共通フィルタ条件では、中項目『Mail』のデータがありません。")
        elif "小項目" not in mail_df.columns:
            st.info("小項目列がないため、小項目別のメルマガ分析を表示できません。")
        else:
            mail_df["小項目"] = mail_df["小項目"].astype(str).str.strip()
            small_options = sorted(
                v for v in mail_df["小項目"].dropna().unique().tolist()
                if v and v.casefold() != "nan"
            )

            # 共通フィルタやファイル変更で候補が変わっても、古い選択値を残さない。
            consume_saved_preference("mail_smalls", small_options, default=small_options[: min(5, len(small_options))], multiple=True)
            saved_smalls = st.session_state.get("mail_smalls", small_options[: min(5, len(small_options))])
            if not isinstance(saved_smalls, list):
                saved_smalls = [saved_smalls]
            saved_smalls = [v for v in saved_smalls if v in small_options]
            if not saved_smalls and small_options:
                saved_smalls = small_options[: min(5, len(small_options))]
            st.session_state["mail_smalls"] = saved_smalls

            selected_smalls = st.multiselect(
                "表示する小項目（複数選択可）",
                small_options,
                key="mail_smalls",
            )

            if not selected_smalls:
                st.info("表示する小項目を1つ以上選択してください。")
            else:
                target_small_df = mail_df[mail_df["小項目"].isin(selected_smalls)].copy()
                campaign_col = "キャンペーン識別" if "キャンペーン識別" in target_small_df.columns else "小項目2"

                if campaign_col not in target_small_df.columns:
                    st.info("キャンペーン識別に使用できる列がありません。")
                else:
                    target_small_df[campaign_col] = target_small_df[campaign_col].astype(str).str.strip()
                    campaigns = sorted(
                        v for v in target_small_df[campaign_col].dropna().unique().tolist()
                        if v and v.casefold() != "nan"
                    )

                    # 小項目を変更するとキャンペーン候補も変わるため、無効な選択値を先に除外する。
                    consume_saved_preference("mail_campaigns", campaigns, default=campaigns[: min(10, len(campaigns))], multiple=True)
                    saved_campaigns = st.session_state.get(
                        "mail_campaigns", campaigns[: min(10, len(campaigns))]
                    )
                    if not isinstance(saved_campaigns, list):
                        saved_campaigns = [saved_campaigns]
                    saved_campaigns = [v for v in saved_campaigns if v in campaigns]
                    if not saved_campaigns and campaigns:
                        saved_campaigns = campaigns[: min(10, len(campaigns))]
                    st.session_state["mail_campaigns"] = saved_campaigns

                    selected_campaigns = st.multiselect(
                        "キャンペーン（複数選択可）",
                        campaigns,
                        key="mail_campaigns",
                    )

                    if not selected_campaigns:
                        st.info("比較するキャンペーンを1つ以上選択してください。")
                    else:
                        target = target_small_df[
                            target_small_df[campaign_col].isin(selected_campaigns)
                        ].copy()

                        if target.empty:
                            st.info("選択条件に該当するデータがありません。")
                        else:
                            mail_res = aggregate(target, ["小項目", campaign_col])
                            available_mail_metrics = [m for m in CORE_METRICS if m in mail_res.columns]

                            if not available_mail_metrics:
                                st.info("集計可能な表示項目がありません。")
                            else:
                                # 一覧表のデフォルト表示は申込件数。
                                consume_saved_preference("mail_metrics", available_mail_metrics, default=[m for m in admin_default("mail_metrics", ["申込件数"]) if m in available_mail_metrics], multiple=True)
                                consume_saved_preference("mail_metric1", available_mail_metrics, default=admin_default("mail_metric1", "申込件数") if admin_default("mail_metric1", "申込件数") in available_mail_metrics else available_mail_metrics[0])
                                consume_saved_preference("mail_metric2", available_mail_metrics, default=admin_default("mail_metric2", "承認率") if admin_default("mail_metric2", "承認率") in available_mail_metrics else available_mail_metrics[0])
                                saved_mail_metrics = st.session_state.get("mail_metrics", ["申込件数"])
                                if not isinstance(saved_mail_metrics, list):
                                    saved_mail_metrics = [saved_mail_metrics]
                                saved_mail_metrics = [m for m in saved_mail_metrics if m in available_mail_metrics]
                                if not saved_mail_metrics:
                                    saved_mail_metrics = (
                                        ["申込件数"] if "申込件数" in available_mail_metrics
                                        else [available_mail_metrics[0]]
                                    )
                                st.session_state["mail_metrics"] = saved_mail_metrics

                                mail_metrics = st.multiselect(
                                    "一覧表の表示項目",
                                    available_mail_metrics,
                                    key="mail_metrics",
                                )

                                default_metric1 = (
                                    "申込件数" if "申込件数" in available_mail_metrics
                                    else available_mail_metrics[0]
                                )
                                default_metric2 = (
                                    "承認率" if "承認率" in available_mail_metrics
                                    else available_mail_metrics[min(1, len(available_mail_metrics) - 1)]
                                )

                                if st.session_state.get("mail_metric1") not in available_mail_metrics:
                                    st.session_state["mail_metric1"] = default_metric1
                                if st.session_state.get("mail_metric2") not in available_mail_metrics:
                                    st.session_state["mail_metric2"] = default_metric2

                                g1, g2 = st.columns(2)
                                metric1 = g1.selectbox(
                                    "比較グラフ1の指標",
                                    available_mail_metrics,
                                    key="mail_metric1",
                                )
                                metric2 = g2.selectbox(
                                    "比較グラフ2の指標",
                                    available_mail_metrics,
                                    key="mail_metric2",
                                )

                                display_cols = list(dict.fromkeys(
                                    ["小項目", campaign_col]
                                    + [m for m in mail_metrics if m in mail_res.columns]
                                ))
                                st.subheader("小項目別一覧")
                                st.dataframe(
                                    format_table(mail_res.loc[:, display_cols]),
                                    width="stretch",
                                    hide_index=True,
                                )

                                st.subheader("小項目・キャンペーン比較グラフ")
                                st.caption(
                                    "1グラフにつき1指標を表示します。横軸は小項目、同じ小項目内でキャンペーンを横並びに比較します。"
                                )

                                # Altairで xOffset を指定し、積み上げではなく集合縦棒に固定する。
                                chart_source = mail_res[["小項目", campaign_col] + list(dict.fromkeys([metric1, metric2]))].copy()
                                chart_source["小項目"] = pd.Categorical(
                                    chart_source["小項目"], categories=selected_smalls, ordered=True
                                )
                                chart_source[campaign_col] = pd.Categorical(
                                    chart_source[campaign_col], categories=selected_campaigns, ordered=True
                                )

                                def draw_grouped_mail_chart(container, metric):
                                    container.markdown(f"#### {metric}")
                                    if metric not in chart_source.columns:
                                        container.info(f"{metric}を集計できません。")
                                        return

                                    plot_df = chart_source[["小項目", campaign_col, metric]].dropna(subset=[metric]).copy()
                                    if plot_df.empty:
                                        container.info("表示できるデータがありません。")
                                        return

                                    y_format = ".1%" if metric in {"承認率", "成約率"} else ",.0f"
                                    tooltip_format = y_format

                                    chart = (
                                        alt.Chart(plot_df)
                                        .mark_bar()
                                        .encode(
                                            x=alt.X(
                                                "小項目:N",
                                                sort=selected_smalls,
                                                title="小項目",
                                                axis=alt.Axis(labelAngle=0),
                                            ),
                                            xOffset=alt.XOffset(
                                                f"{campaign_col}:N",
                                                sort=selected_campaigns,
                                            ),
                                            y=alt.Y(
                                                f"{metric}:Q",
                                                title=metric,
                                                axis=alt.Axis(format=y_format),
                                                stack=None,
                                            ),
                                            color=alt.Color(
                                                f"{campaign_col}:N",
                                                sort=selected_campaigns,
                                                title="キャンペーン",
                                            ),
                                            tooltip=[
                                                alt.Tooltip("小項目:N", title="小項目"),
                                                alt.Tooltip(f"{campaign_col}:N", title="キャンペーン"),
                                                alt.Tooltip(f"{metric}:Q", title=metric, format=tooltip_format),
                                            ],
                                        )
                                        .properties(height=360)
                                    )
                                    container.altair_chart(chart, width="stretch")

                                chart_left, chart_right = st.columns(2)
                                draw_grouped_mail_chart(chart_left, metric1)
                                draw_grouped_mail_chart(chart_right, metric2)

                                st.download_button(
                                    "📥 メルマガ分析をダウンロード",
                                    to_excel(mail_res),
                                    "メルマガ分析_小項目別.xlsx",
                                )

                                # ======================
                                # メルマガ × セグメント分析
                                # ======================
                                st.divider()
                                st.subheader("メルマガのセグメント別分析")
                                st.caption(
                                    "セグメント別タブと同じ分析項目から1つ選び、現在選択中の小項目・キャンペーンを比較します。"
                                )

                                mail_segment_options = [
                                    c for c in dimension_cols
                                    if c in target.columns
                                    and c not in {"小項目", campaign_col}
                                ]

                                if not mail_segment_options:
                                    st.info("セグメント分析に使用できる項目がありません。")
                                else:
                                    default_mail_segment = (
                                        "年齢グループ"
                                        if "年齢グループ" in mail_segment_options
                                        else mail_segment_options[0]
                                    )
                                    consume_saved_preference("mail_segment_col", mail_segment_options, default=default_mail_segment)
                                    consume_saved_preference("mail_segment_metric", available_mail_metrics, default=default_metric1)
                                    if st.session_state.get("mail_segment_col") not in mail_segment_options:
                                        st.session_state["mail_segment_col"] = default_mail_segment

                                    mail_segment_col = st.selectbox(
                                        "分析項目",
                                        mail_segment_options,
                                        key="mail_segment_col",
                                    )

                                    if st.session_state.get("mail_segment_metric") not in available_mail_metrics:
                                        st.session_state["mail_segment_metric"] = default_metric1

                                    mail_segment_metric = st.selectbox(
                                        "セグメント比較グラフの指標",
                                        available_mail_metrics,
                                        key="mail_segment_metric",
                                    )

                                    segment_source = target.copy()
                                    segment_source = segment_source[
                                        segment_source[mail_segment_col].notna()
                                    ].copy()
                                    segment_source[mail_segment_col] = (
                                        segment_source[mail_segment_col]
                                        .astype(str)
                                        .str.strip()
                                    )
                                    segment_source = segment_source[
                                        segment_source[mail_segment_col].ne("")
                                        & segment_source[mail_segment_col].str.casefold().ne("nan")
                                    ]

                                    if segment_source.empty:
                                        st.info("選択した分析項目に表示可能なデータがありません。")
                                    else:
                                        mail_segment_res = aggregate(
                                            segment_source,
                                            [mail_segment_col, "小項目", campaign_col],
                                        )

                                        segment_display_metrics = [
                                            m for m in mail_metrics
                                            if m in mail_segment_res.columns
                                        ]
                                        if not segment_display_metrics:
                                            segment_display_metrics = [mail_segment_metric]

                                        segment_display_cols = list(dict.fromkeys(
                                            [mail_segment_col, "小項目", campaign_col]
                                            + segment_display_metrics
                                        ))

                                        st.markdown("#### セグメント別一覧")
                                        st.dataframe(
                                            format_table(
                                                mail_segment_res.loc[:, segment_display_cols]
                                            ),
                                            width="stretch",
                                            hide_index=True,
                                        )

                                        st.markdown(f"#### {mail_segment_col}別：{mail_segment_metric}")
                                        segment_chart_df = mail_segment_res[
                                            [mail_segment_col, "小項目", campaign_col, mail_segment_metric]
                                        ].dropna(subset=[mail_segment_metric]).copy()

                                        if segment_chart_df.empty:
                                            st.info("グラフに表示できるデータがありません。")
                                        else:
                                            segment_chart_df["比較系列"] = (
                                                segment_chart_df["小項目"].astype(str)
                                                + "｜"
                                                + segment_chart_df[campaign_col].astype(str)
                                            )
                                            segment_order = list(dict.fromkeys(
                                                segment_chart_df[mail_segment_col].astype(str).tolist()
                                            ))
                                            series_order = [
                                                f"{small}｜{campaign}"
                                                for small in selected_smalls
                                                for campaign in selected_campaigns
                                                if f"{small}｜{campaign}"
                                                in set(segment_chart_df["比較系列"].tolist())
                                            ]

                                            y_format = (
                                                ".1%"
                                                if mail_segment_metric in {"承認率", "成約率"}
                                                else ",.0f"
                                            )
                                            segment_chart = (
                                                alt.Chart(segment_chart_df)
                                                .mark_bar()
                                                .encode(
                                                    x=alt.X(
                                                        f"{mail_segment_col}:N",
                                                        sort=segment_order,
                                                        title=mail_segment_col,
                                                        axis=alt.Axis(labelAngle=-30),
                                                    ),
                                                    xOffset=alt.XOffset(
                                                        "比較系列:N",
                                                        sort=series_order,
                                                    ),
                                                    y=alt.Y(
                                                        f"{mail_segment_metric}:Q",
                                                        title=mail_segment_metric,
                                                        axis=alt.Axis(format=y_format),
                                                        stack=None,
                                                    ),
                                                    color=alt.Color(
                                                        "比較系列:N",
                                                        sort=series_order,
                                                        title="小項目｜キャンペーン",
                                                    ),
                                                    tooltip=[
                                                        alt.Tooltip(
                                                            f"{mail_segment_col}:N",
                                                            title=mail_segment_col,
                                                        ),
                                                        alt.Tooltip("小項目:N", title="小項目"),
                                                        alt.Tooltip(
                                                            f"{campaign_col}:N",
                                                            title="キャンペーン",
                                                        ),
                                                        alt.Tooltip(
                                                            f"{mail_segment_metric}:Q",
                                                            title=mail_segment_metric,
                                                            format=y_format,
                                                        ),
                                                    ],
                                                )
                                                .properties(height=420)
                                            )
                                            st.altair_chart(
                                                segment_chart,
                                                width="stretch",
                                            )

                                        st.download_button(
                                            "📥 メルマガのセグメント分析をダウンロード",
                                            to_excel(mail_segment_res),
                                            "メルマガ分析_セグメント別.xlsx",
                                        )

# ======================
# ブラウザ設定の保存・初期化
# ======================
st.sidebar.divider()
st.sidebar.subheader("⚙️ 個人設定")
if LocalStorage is None:
    st.sidebar.warning("ブラウザ保存機能を使うには requirements.txt に streamlit-local-storage を追加してください。")
else:
    st.sidebar.caption("コード側のデフォルトを維持し、保存した項目だけこのブラウザで復元します。")
    if st.sidebar.button("💾 現在の設定を保存", width="stretch", key="save_browser_preferences"):
        payload = json.dumps({"schema_version": PREFERENCE_SCHEMA_VERSION, "values": collect_preferences()}, ensure_ascii=False)
        storage_set_item(local_storage, PREFERENCE_STORAGE_KEY, payload, component_key="save_preferences_component")
        time.sleep(0.8)
        st.sidebar.success("このブラウザに設定を保存しました。")

    if st.sidebar.button("↩️ 保存設定を初期化", width="stretch", key="reset_browser_preferences"):
        storage_set_item(local_storage, PREFERENCE_STORAGE_KEY, "", component_key="reset_preferences_component")
        for pref_key in PREFERENCE_KEYS:
            st.session_state.pop(pref_key, None)
        st.session_state.pop("_pending_browser_preferences", None)
        st.session_state["_browser_preferences_loaded"] = True
        time.sleep(0.8)
        st.rerun()
